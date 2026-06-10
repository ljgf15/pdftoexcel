# -*- coding: utf-8 -*-
"""
PDF采购订单提取核心逻辑。

当前可选字段已限制为 ALLOWED_COLUMNS 中的字段。
支持：
1. 天吴采购单块状结构。
2. KONE Pos 明细结构。
3. KONE 门类 HH/LL 尺寸字段。
4. KONE 木箱 DIM_CAR_BOX_INNER_* 尺寸字段。
"""

import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# 1. 在 ALLOWED_COLUMNS 里新增一列
ALLOWED_COLUMNS = [
    "送货日期",
    "订单号",
    "件号",
    "物料名称",
    "物料规格",
    "单位",
    "数量",
    "未税单价",
    "未税总价",
    "印刷",
    "Pos",
    "Material",
    "Quantity",
    "Unit",
    "Price",
    "Amount",
    "Sales order ref",
    "Sales order no",
    "Sales order item",
    "轿厢净开门宽度_LL_mm",
    "轿门净高_HH_mm",
    "门尺寸_LL*HH",
    "DIM_CAR_BOX_INNER_LENGTH_mm",
    "DIM_CAR_BOX_INNER_WIDTH_mm",
    "DIM_CAR_BOX_INNER_HEIGHT_mm",
    "长宽高",
    "包装箱类别",
]


ALIASES = {
    "长*宽*高": "长宽高",
    "L*W*H": "长宽高",
    "尺寸": "长宽高",
    "规格": "物料规格",
    "物料号": "Material",
    "料号": "件号",
    "订单编号": "订单号",
    "采购单号": "订单号",
    "包装箱类型": "包装箱类别",
    "包装类型": "包装箱类别",
    "交货日": "送货日期",
    "交货日期": "送货日期",
}

# 2. 新增这个函数，放在 extractor.py 任意函数区即可
def extract_packaging_box_type(block):
    """
    提取包装箱类型

    PDF示例：

    包装箱类型
    1      1,普通

    返回：
    1,普通
    """

    if not block:
        return ""

    text = normalize_text(block)

    patterns = [
        # 包装箱类型 1 1,普通
        r"包装箱类型\s+\d+\s+(\d+\s*[,，]\s*[^\s\n]+)",

        # 包装箱类型: 1,普通
        r"包装箱类型\s*[:：]?\s*(\d+\s*[,，]\s*[^\s\n]+)",

        # 跨行情况
        r"包装箱类型.*?\n.*?(\d+\s*[,，]\s*[^\s\n]+)",
    ]

    for pattern in patterns:
        m = re.search(pattern, text, re.I | re.S)
        if m:
            return (
                m.group(1)
                .replace("，", ",")
                .replace(" ", "")
                .strip()
            )

    # 紧凑文本兜底
    compact = compact_text(text)

    m = re.search(
        r"包装箱类型\d+(\d+[,，][^\s]+)",
        compact,
        re.I
    )

    if m:
        return (
            m.group(1)
            .replace("，", ",")
            .strip()
        )

    return ""



def canonical_column_name(name: str) -> str:
    return ALIASES.get(str(name).strip(), str(name).strip())


def clean_text(value):
    return str(value or "").replace("\xa0", " ").strip()


def normalize_lines(text):
    lines = []
    for raw in (text or "").replace("\xa0", " ").replace("\r", "\n").split("\n"):
        line = re.sub(r"[ \t]+", " ", raw).strip()
        if line:
            lines.append(line)
    return lines


def normalize_text(text):
    return "\n".join(normalize_lines(text))


def flatten_text(text):
    return " ".join(normalize_lines(text))


def compact_text(text):
    return re.sub(r"\s+", "", str(text or ""))


def clean_number(value):
    if value is None:
        return ""
    text = str(value).strip().replace(",", "")
    if not text:
        return ""
    try:
        number = float(text)
        if number.is_integer():
            return int(number)
        return number
    except Exception:
        return text


def as_plain_number_text(value):
    value = clean_number(value)
    if value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def split_quantity_unit(value):
    raw = str(value or "").strip().replace(" ", "")
    if not raw:
        return "", "", ""
    match = re.match(r"^([\d,]+(?:\.\d+)?)([A-Za-z]+|[\u4e00-\u9fa5]+)?$", raw)
    if not match:
        return raw, "", raw
    return clean_number(match.group(1)), match.group(2) or "", raw


def normalize_sales_order_ref(value):
    """统一 Sales order ref 的写法，保留前导 0。"""
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = raw.replace("／", "/").replace("\\", "/").replace(",", "")
    raw = re.sub(r"\s*/\s*", "/", raw)
    raw = re.sub(r"\s+", "", raw)
    raw = raw.strip(" .:：;，,")
    return raw


def split_sales_order_ref(value):
    raw = normalize_sales_order_ref(value)
    if not raw:
        return "", "", ""
    parts = raw.split("/", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip(), raw
    return raw, "", raw


def extract_sales_order_ref(block):
    """
    提取英文 KONE PDF 中的 Sales order ref。
    支持：
    Sales order ref. 1234567890/10
    Sales order ref. Z350920443/6000
    Sales order ref: 1234567890 / 000010
    Sales order reference 1234567890 / 10
    Sales order ref. 1234567890 Sales order item 10
    Sales order ref. Z350920443 Sales order item 6000
    Sales order ref. 1234567890
    """
    text = str(block or "")
    if not text:
        return ""

    # 订单号可能是纯数字，也可能带字母前缀，例如 Z350920443/6000。
    token = r"[A-Z0-9][A-Z0-9,]{2,30}"
    item_token = r"[A-Z0-9][A-Z0-9,]{0,12}"
    label = r"Sales\s+order\s+ref(?:erence)?\s*\.?\s*[:：]?\s*"

    def is_valid_sales_order_candidate(candidate):
        candidate = normalize_sales_order_ref(candidate)
        if not re.search(r"\d", candidate):
            return False
        # 避免把后面的 Project ref / Purchase order 等标签误拼进来。
        bad_prefixes = ("PROJECTREF", "PURCHASEORDER", "MATERIAL", "DESCRIPTION", "DELIVERY")
        if candidate.upper().startswith(bad_prefixes):
            return False
        return bool(
            re.fullmatch(r"[0-9A-Za-z]+/[0-9A-Za-z]+", candidate)
            or re.fullmatch(r"[0-9A-Za-z]{6,}", candidate)
        )

    patterns = [
        # 标准写法：Sales order ref. 123/10 或 Sales order ref. Z350920443/6000。
        rf"{label}({token}\s*[/／]\s*{item_token})(?=[^A-Z0-9]|$)",
        # 分开写：Sales order ref. 1234567890 Sales order item 10。
        rf"{label}({token})\s+Sales\s+order\s+item\.?\s*[:：]?\s*({item_token})(?=[^A-Z0-9]|$)",
        # 简化写法：Sales order ref. 1234567890 item 10 / position 10 / line 10。
        rf"{label}({token})\s+(?:item|position|pos\.?|line)\s*[:：]?\s*({item_token})(?=[^A-Z0-9]|$)",
        # 只有订单号，没有 item 的写法。
        rf"{label}({token})(?=[^A-Z0-9]|$)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if not match:
            continue
        if len(match.groups()) >= 2 and match.group(2):
            candidate = f"{match.group(1)}/{match.group(2)}"
        else:
            candidate = match.group(1)
        candidate = normalize_sales_order_ref(candidate)
        if is_valid_sales_order_candidate(candidate):
            return candidate

    # 紧凑文本兜底：Salesorderref.Z350920443/6000。
    compact = compact_text(text).replace("／", "/")
    match = re.search(r"Salesorderref(?:erence)?\.?[:：]?([A-Z0-9][A-Z0-9,]{2,30}/[0-9][0-9,]{0,12})(?=[^0-9]|$)", compact, re.I)
    if match:
        candidate = normalize_sales_order_ref(match.group(1))
        if is_valid_sales_order_candidate(candidate):
            return candidate

    match = re.search(r"Salesorderref(?:erence)?\.?[:：]?([A-Z0-9][A-Z0-9,]{5,30})(?=[^A-Z0-9]|$)", compact, re.I)
    if match:
        candidate = normalize_sales_order_ref(match.group(1))
        if is_valid_sales_order_candidate(candidate):
            return candidate

    return ""


def find_one(pattern, text, default="", flags=re.I):
    match = re.search(pattern, text or "", flags)
    if not match:
        return default
    return match.group(1).strip()


def find_money_after_label(label, text):
    pattern = rf"{re.escape(label)}\s*[:：]?\s*([0-9,]+(?:\.\d+)?)"
    return clean_number(find_one(pattern, text, flags=re.I))


def format_lwh(length, width, height):
    values = [as_plain_number_text(length), as_plain_number_text(width), as_plain_number_text(height)]
    if any(v == "" for v in values):
        return ""
    return "*".join(values)


def format_ll_hh(ll, hh):
    ll_text = as_plain_number_text(ll)
    hh_text = as_plain_number_text(hh)
    if not ll_text or not hh_text:
        return ""
    return f"{ll_text}*{hh_text}"


def extract_text_from_pdf(pdf_path):
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            parts.append(f"\n---PAGE {page_no}---\n{text}")
    return "\n".join(parts)


def parse_pdf(pdf_path):
    full_text = extract_text_from_pdf(pdf_path)
    rows = []

    tianwu_rows = parse_tianwu_pdf(pdf_path, full_text)
    kone_rows = parse_kone_pdf(pdf_path, full_text)

    rows.extend(tianwu_rows)
    rows.extend(kone_rows)

    # 如果同一份 PDF 被两个解析器同时命中，保留全部；实际格式通常只命中一种。
    return rows


def parse_tianwu_pdf(pdf_path, full_text):
    flat = flatten_text(full_text)
    if not any(key in flat for key in ["采购单号", "料号", "参数及备注", "未税单价", "采购量"]):
        return []

    order_no = find_one(r"采购单号\s*[:：]?\s*([A-Z0-9_\-]+)", flat, flags=re.I)
    if not order_no:
        order_no = find_one(r"\b(PN[0-9A-Z_\-]+)\b", flat, flags=re.I)

    supplier_print = extract_printing(flat)

    # 块起点：料号:
    starts = list(re.finditer(r"料号\s*[:：]\s*([A-Z0-9_\-]+)", flat, flags=re.I))
    rows = []

    for idx, match in enumerate(starts):
        start = match.start()
        end = starts[idx + 1].start() if idx + 1 < len(starts) else len(flat)
        block = flat[start:end]
        item_no = match.group(1).strip()

        qty_raw = find_one(r"采购量\s*[:：]\s*([0-9,]+(?:\.\d+)?\s*[A-Za-z\u4e00-\u9fa5]+)", block, flags=re.I)
        qty, unit, _ = split_quantity_unit(qty_raw)

        remark = find_one(
            r"参数及备注\s*[:：]\s*(.*?)(?:未税单价|含税单价|采购量|交货日|仓库库位|$)",
            block,
            flags=re.I,
        )

        material_name = extract_tianwu_material_name(block, remark)
        material_spec = extract_lwh_from_remark(remark) or extract_lwh_from_remark(block)

        row = {
            "送货日期": find_one(r"交货日\s*[:：]\s*([0-9]{4}[/-][0-9]{1,2}[/-][0-9]{1,2})", block, flags=re.I),
            "订单号": order_no,
            "件号": item_no,
            "物料名称": material_name,
            "物料规格": material_spec,
            "单位": unit,
            "数量": qty,
            "未税单价": find_money_after_label("未税单价", block),
            "未税总价": find_money_after_label("未税金额", block) or find_money_after_label("未税总价", block),
            "印刷": extract_printing(remark) or supplier_print,
        }
        rows.append(row)

    return rows


def extract_tianwu_material_name(block, remark):
    for label in ["品名规格", "物料名称", "品名", "名称"]:
        value = find_one(
            rf"{label}\s*[:：]\s*(.*?)(?:识别号|图号|参数及备注|未税单价|含税单价|采购量|交货日|$)",
            block,
            flags=re.I,
        )
        value = clean_material_name(value)
        if value:
            return value

    text = f"{block} {remark}"
    if "板条箱" in text:
        return "板条箱"
    if "木箱" in text and "非加固" in text:
        return "普通木箱"
    if "木箱" in text:
        return "木箱"
    if "纸箱" in text:
        return "纸箱"
    if "托盘" in text:
        return "托盘"
    if "包装箱" in text:
        return "包装箱"
    return ""


def clean_material_name(value):
    value = clean_text(value)
    if not value:
        return ""
    value = re.split(r"(识别号|图号|参数及备注|未税单价|含税单价|采购量|交货日|仓库库位)", value)[0]
    value = re.sub(r"[\^;]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip(" ：:;，,")
    return value[:80].strip()


def extract_lwh_from_remark(text):
    if not text:
        return ""

    # ^L=2050;W=1850;H=800
    match = re.search(
        r"\bL\s*=\s*([0-9,]+(?:\.\d+)?)\s*;?\s*W\s*=\s*([0-9,]+(?:\.\d+)?)\s*;?\s*H\s*=\s*([0-9,]+(?:\.\d+)?)",
        text,
        re.I,
    )
    if match:
        return format_lwh(match.group(1), match.group(2), match.group(3))

    # 2050*1850*800
    match = re.search(r"([0-9,]+(?:\.\d+)?)\s*[*xX×]\s*([0-9,]+(?:\.\d+)?)\s*[*xX×]\s*([0-9,]+(?:\.\d+)?)", text)
    if match:
        return format_lwh(match.group(1), match.group(2), match.group(3))

    return ""


def extract_printing(text):
    if not text:
        return ""

    # 参数及备注: ^TKE;...
    match = re.search(r"\^([A-Z]{2,10})\s*;", text)
    if match:
        return match.group(1).strip()

    match = re.search(r"印刷\s*[:：]?\s*([A-Z0-9_\-]+)", text, re.I)
    if match:
        return match.group(1).strip()

    for brand in ["TKE", "KONE", "OTIS", "TK", "蒂升"]:
        if re.search(rf"\b{re.escape(brand)}\b", text, re.I):
            return brand
    return ""


def parse_kone_pdf(pdf_path, full_text):
    flat = flatten_text(full_text)

    # KONE 明细行，例如：
    # 10 KM52059574V003 23.05.2026 1PC 180.00 180.00
    item_pattern = re.compile(
        r"(?m)(?:^|\n|\s)"
        r"(\d{1,5})\s+"
        r"([A-Z0-9]{6,}(?:V\d{3,4})?)\s+"
        r"(\d{2}\.\d{2}\.\d{4})\s+"
        r"([0-9,]+(?:\.\d+)?\s*[A-Za-z]+)\s+"
        r"([0-9,]+(?:\.\d{2}))\s+"
        r"([0-9,]+(?:\.\d{2}))"
    )

    starts = list(item_pattern.finditer(flat))
    if not starts:
        return []

    po_no = find_one(r"Purchase\s+order\s+No\.\s*([0-9]{6,})", flat, flags=re.I)
    if not po_no:
        po_no = find_one(r"\bNo\.\s*([0-9]{6,})\b", flat, flags=re.I)

    rows = []
    for idx, match in enumerate(starts):
        start = match.start()
        end = starts[idx + 1].start() if idx + 1 < len(starts) else len(flat)
        block = flat[start:end]

        qty, unit, _ = split_quantity_unit(match.group(4))
        sales_ref = extract_sales_order_ref(block)
        sales_no, sales_item, _ = split_sales_order_ref(sales_ref)

        hh = extract_hh_mm(block)
        ll = extract_ll_mm(block)

        height = extract_dim_value(block, "HEIGHT")
        length = extract_dim_value(block, "LENGTH")
        width = extract_dim_value(block, "WIDTH")

        row = {
            "送货日期": match.group(3),
            "订单号": po_no,
            "件号": match.group(2),
            "物料名称": extract_kone_material_name(block, match.group(2)),
            "物料规格": "",
            "单位": unit,
            "数量": qty,
            "未税单价": clean_number(match.group(5)),
            "未税总价": clean_number(match.group(6)),
            "印刷": extract_printing(block),
            "Pos": clean_number(match.group(1)),
            "Material": match.group(2),
            "Quantity": qty,
            "Unit": unit,
            "Price": clean_number(match.group(5)),
            "Amount": clean_number(match.group(6)),
            "Sales order ref": sales_ref,
            "Sales order no": sales_no,
            "Sales order item": sales_item,
            "轿厢净开门宽度_LL_mm": ll,
            "轿门净高_HH_mm": hh,
            "门尺寸_LL*HH": format_ll_hh(ll, hh),
            "DIM_CAR_BOX_INNER_LENGTH_mm": length,
            "DIM_CAR_BOX_INNER_WIDTH_mm": width,
            "DIM_CAR_BOX_INNER_HEIGHT_mm": height,
            "长宽高": format_lwh(length, width, height),
            "包装箱类别": extract_packaging_box_type(block),
        }

        # 物料规格兜底：门类用 LL*HH，木箱用长宽高。
        row["物料规格"] = row["门尺寸_LL*HH"] or row["长宽高"]

        rows.append(row)

    return rows


def extract_package_box_type(block):
    """
    提取 KONE 明细中的包装箱类别/包装箱类型。
    示例：
    包装箱类型 1 1,普通
    包装箱类型 1,普通
    返回：普通
    """
    text = str(block or "")
    if not text:
        return ""

    # 常见格式：标签后可能先有一个参数值 0/1，再出现 “1,普通”。
    patterns = [
        r"包装箱(?:类别|类型)\s*[:：]?\s*(?:[0-9]+\s+)?[0-9]+\s*[,，]\s*([^\s,，;；]+)",
        r"包装箱(?:类别|类型)\s*[:：]?\s*([^\s,，;；]+)",
        r"Package\s*(?:box\s*)?(?:type|category)\s*[:：]?\s*(?:[0-9]+\s+)?[0-9]+\s*[,，]\s*([^\s,，;；]+)",
    ]
    for pattern in patterns:
        value = find_one(pattern, text, flags=re.I)
        value = clean_text(value).strip(" ,，;；")
        if value and not re.fullmatch(r"[0-9]+", value):
            return value

    # 紧凑文本兜底：包装箱类型1,普通。
    compact = compact_text(text)
    value = find_one(r"包装箱(?:类别|类型)(?:[0-9]+)?[0-9][,，]([^,，;；\s]+)", compact, flags=re.I)
    if value:
        return clean_text(value).strip(" ,，;；")

    return ""

def extract_dim_value(block, kind):
    """
    提取 DIM_CAR_BOX_INNER_HEIGHT/LENGTH/WIDTH 的值。
    支持：
    DIM_CAR_BOX_INNER_HEIGHT 520 mm
    DIM CAR BOX INNER HEIGHT 520 mm
    DIM_CAR_BOX_INNER_HEIGHT_mm 520
    标签和值被空格拆开的情况。
    """
    kind = kind.upper()
    patterns = [
        rf"DIM[_\s]*CAR[_\s]*BOX[_\s]*INNER[_\s]*{kind}(?:[_\s]*mm)?\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*(?:mm)?",
        rf"DIM_CAR_BOX_INNER_{kind}\s+([0-9,]+(?:\.\d+)?)\s*mm",
        rf"{kind}\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*mm",
    ]

    for pattern in patterns:
        value = find_one(pattern, block, flags=re.I)
        if value:
            return clean_number(value)

    # 紧凑文本兜底
    compact = compact_text(block)
    pattern = rf"DIMCARBOXINNER{kind}(?:MM)?([0-9,]+(?:\.\d+)?)MM?"
    value = find_one(pattern, compact, flags=re.I)
    if value:
        return clean_number(value)

    return ""


def extract_hh_mm(block):
    patterns = [
        r"轿门净高\s*[,，]?\s*HH\.?\s*\[?mm\]?\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*mm",
        r"轿门净高\s*HH\s*\(?mm\)?\s*([0-9,]+(?:\.\d+)?)",
        r"HH\.?\s*\[?mm\]?\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*mm",
        r"CAR\s+DOOR\s+HEIGHT.*?([0-9,]+(?:\.\d+)?)\s*mm",
    ]
    for pattern in patterns:
        value = find_one(pattern, block, flags=re.I)
        if value:
            return clean_number(value)

    compact = compact_text(block)
    value = find_one(r"轿门净高HH(?:MM)?([0-9,]+(?:\.\d+)?)MM?", compact, flags=re.I)
    if value:
        return clean_number(value)
    return ""


def extract_ll_mm(block):
    patterns = [
        r"轿厢净开门宽度\s*[,，]?\s*LL\.?\s*\(?\[?mm\]?\)?\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*mm",
        r"轿厢净开门宽度\s*LL\s*\(?mm\)?\s*([0-9,]+(?:\.\d+)?)",
        r"LL\.?\s*\[?mm\]?\s*[:：]?\s*([0-9,]+(?:\.\d+)?)\s*mm",
        r"CAR\s+DOOR\s+WIDTH.*?([0-9,]+(?:\.\d+)?)\s*mm",
    ]
    for pattern in patterns:
        value = find_one(pattern, block, flags=re.I)
        if value:
            return clean_number(value)

    compact = compact_text(block)
    value = find_one(r"轿厢净开门宽度LL(?:MM)?([0-9,]+(?:\.\d+)?)MM?", compact, flags=re.I)
    if value:
        return clean_number(value)
    return ""


def extract_kone_material_name(block, material_code):
    # 先查明确描述标签
    for label in ["Material description", "Item description", "Description"]:
        value = find_one(rf"{label}\s*[:：]?\s*([^\n]+)", block, flags=re.I)
        value = clean_material_name(value)
        if value:
            return value

    # 从块里推断包装类型
    if "板条箱" in block:
        return "板条箱"
    if "木箱" in block and "非加固" in block:
        return "普通木箱"
    if "木箱" in block:
        return "木箱"
    if "纸箱" in block:
        return "纸箱"
    if "包装箱" in block:
        return "包装箱"

    # 从物料代码后面的短描述尝试取
    idx = block.find(material_code)
    if idx >= 0:
        tail = block[idx + len(material_code): idx + len(material_code) + 160]
        tail = re.sub(r"\d{2}\.\d{2}\.\d{4}.*", "", tail).strip()
        tail = clean_material_name(tail)
        if tail and not re.fullmatch(r"[0-9A-Za-z .,\-/]+", tail):
            return tail

    return ""


def get_row_value(row, column):
    column = canonical_column_name(column)
    if column not in ALLOWED_COLUMNS:
        return ""
    return row.get(column, "")


def write_excel(rows, columns, output_path, errors=None):
    errors = errors or []
    columns = [canonical_column_name(c) for c in columns if canonical_column_name(c) in ALLOWED_COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = "提取结果"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_index, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        for col_index, name in enumerate(columns, start=1):
            value = get_row_value(row, name)
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    if rows and columns:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    for col_index, name in enumerate(columns, start=1):
        max_len = len(str(name))
        for row_index in range(2, min(len(rows) + 2, 300)):
            value = ws.cell(row=row_index, column=col_index).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(col_index)].width = max(10, min(max_len + 4, 60))

    for col_index, name in enumerate(columns, start=1):
        if name in {"未税单价", "未税总价", "Price", "Amount"}:
            for row_index in range(2, len(rows) + 2):
                ws.cell(row=row_index, column=col_index).number_format = "#,##0.00"

    log_ws = wb.create_sheet("错误日志")
    log_headers = ["PDF文件", "状态/错误"]
    for col_index, name in enumerate(log_headers, start=1):
        cell = log_ws.cell(row=1, column=col_index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    if errors:
        for row_index, error in enumerate(errors, start=2):
            log_ws.cell(row=row_index, column=1, value=error.get("file", "")).border = border
            log_ws.cell(row=row_index, column=2, value=error.get("error", "")).border = border
    else:
        log_ws.cell(row=2, column=1, value="无").border = border
        log_ws.cell(row=2, column=2, value="全部PDF处理成功").border = border

    log_ws.column_dimensions["A"].width = 45
    log_ws.column_dimensions["B"].width = 100

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
